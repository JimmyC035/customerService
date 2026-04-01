import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

from app.constants import COLORS, COLS

# 讓 matplotlib 顯示中文
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'Microsoft JhengHei', 'SimHei', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False


class AnalyticsPanel:
    """銷售分析分頁"""

    def __init__(self, parent, get_df):
        """
        parent: 分頁容器
        get_df: callable，回傳目前的 DataFrame
        """
        self.parent = parent
        self.get_df = get_df

        self._build_controls()
        self._build_summary()
        self._build_charts()

    def _build_controls(self):
        ctrl = tk.Frame(self.parent, bg=COLORS["card"],
                        highlightbackground=COLORS["border"], highlightthickness=1)
        ctrl.pack(fill="x", padx=24, pady=(16, 8))

        inner = tk.Frame(ctrl, bg=COLORS["card"])
        inner.pack(fill="x", padx=20, pady=12)

        tk.Label(inner, text="起始月份:", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Arial", 11)).pack(side="left")
        self.start_month = ttk.Combobox(inner, width=10, font=("Arial", 11))
        self.start_month.pack(side="left", padx=(4, 16))

        tk.Label(inner, text="結束月份:", bg=COLORS["card"], fg=COLORS["text"],
                 font=("Arial", 11)).pack(side="left")
        self.end_month = ttk.Combobox(inner, width=10, font=("Arial", 11))
        self.end_month.pack(side="left", padx=(4, 16))

        tk.Button(inner, text="查詢", command=self.refresh,
                  bg=COLORS["primary"], fg="white",
                  font=("Arial", 11, "bold"), relief="flat",
                  padx=16, pady=4, cursor="hand2").pack(side="left", padx=4)

        tk.Button(inner, text="📊 匯出報表", command=self.export_report,
                  bg=COLORS["success"], fg="white",
                  font=("Arial", 11, "bold"), relief="flat",
                  padx=16, pady=4, cursor="hand2").pack(side="right", padx=4)

    def _build_summary(self):
        self.summary_frame = tk.Frame(self.parent, bg=COLORS["bg"])
        self.summary_frame.pack(fill="x", padx=24, pady=(0, 8))

        self.summary_cards = {}
        labels = [
            ("總營收", "total_revenue"),
            ("總利潤", "total_profit"),
            ("利潤率", "profit_rate"),
            ("訂單數", "order_count"),
        ]
        for text, key in labels:
            card = tk.Frame(self.summary_frame, bg=COLORS["card"],
                            highlightbackground=COLORS["border"], highlightthickness=1)
            card.pack(side="left", fill="x", expand=True, padx=4)

            tk.Label(card, text=text, bg=COLORS["card"], fg=COLORS["text_light"],
                     font=("Arial", 10)).pack(anchor="w", padx=16, pady=(12, 0))
            val_label = tk.Label(card, text="--", bg=COLORS["card"], fg=COLORS["text"],
                                 font=("Arial", 18, "bold"))
            val_label.pack(anchor="w", padx=16, pady=(0, 12))
            self.summary_cards[key] = val_label

    def _build_charts(self):
        chart_frame = tk.Frame(self.parent, bg=COLORS["card"],
                               highlightbackground=COLORS["border"], highlightthickness=1)
        chart_frame.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        self.fig = Figure(figsize=(12, 5), dpi=100, facecolor='white')
        self.ax_bar = self.fig.add_subplot(121)
        self.ax_line = self.fig.add_subplot(122)
        self.fig.subplots_adjust(wspace=0.35, left=0.06, right=0.96, top=0.9, bottom=0.18)

        self.canvas = FigureCanvasTkAgg(self.fig, master=chart_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=8)

    # ─── 資料處理 ───

    def _prepare_df(self):
        """取得 DataFrame 並轉換數值欄位"""
        df = self.get_df().copy()
        if df.empty:
            return df

        df["日期"] = df["日期"].astype(str).str.split(" ").str[0]
        for col in ["總價", "利潤", "實拿", "數量"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        if "品項" in df.columns:
            df["品項"] = df["品項"].astype(str)

        df["月份"] = df["日期"].str[:7]
        # 過濾掉無效日期（空字串、不符格式的）
        df = df[df["月份"].str.match(r'^\d{4}-\d{2}$', na=False)]
        return df

    def _filter_by_range(self, df):
        """依日期範圍篩選"""
        start = self.start_month.get().strip()
        end = self.end_month.get().strip()
        if start:
            df = df[df["月份"] >= start]
        if end:
            df = df[df["月份"] <= end]
        return df

    def _populate_months(self, df):
        """更新月份下拉選單"""
        if df.empty:
            return
        months = sorted([str(m) for m in df["月份"].unique() if m])
        self.start_month['values'] = months
        self.end_month['values'] = months
        if not self.start_month.get() and months:
            self.start_month.set(months[0])
        if not self.end_month.get() and months:
            self.end_month.set(months[-1])

    # ─── 刷新 ───

    def refresh(self):
        df = self._prepare_df()
        if df.empty:
            return

        self._populate_months(df)
        filtered = self._filter_by_range(df)

        self._update_summary(filtered)
        self._draw_product_ranking(filtered)
        self._draw_monthly_trend(filtered)
        self.canvas.draw()

    def _update_summary(self, df):
        total_rev = df["總價"].sum()
        total_profit = df["利潤"].sum()
        rate = (total_profit / total_rev * 100) if total_rev > 0 else 0
        count = len(df)

        self.summary_cards["total_revenue"].config(
            text=f"${total_rev:,.0f}")
        self.summary_cards["total_profit"].config(
            text=f"${total_profit:,.0f}")
        self.summary_cards["profit_rate"].config(
            text=f"{rate:.1f}%")
        self.summary_cards["order_count"].config(
            text=str(count))

    def _draw_product_ranking(self, df):
        self.ax_bar.clear()
        if df.empty or "品項" not in df.columns:
            return

        ranking = df.groupby("品項")["總價"].sum().sort_values(ascending=True).tail(10)

        colors = [COLORS["primary"]] * len(ranking)
        bars = self.ax_bar.barh(ranking.index, ranking.values, color=colors, height=0.6)
        self.ax_bar.set_title("品項銷售排行 Top 10", fontsize=12, fontweight='bold', pad=10)
        self.ax_bar.set_xlabel("營收")
        self.ax_bar.tick_params(axis='y', labelsize=9)

        # 在 bar 上顯示數值
        for bar, val in zip(bars, ranking.values):
            self.ax_bar.text(bar.get_width() + ranking.max() * 0.01,
                             bar.get_y() + bar.get_height() / 2,
                             f'${val:,.0f}', va='center', fontsize=8)

    def _draw_monthly_trend(self, df):
        self.ax_line.clear()
        if df.empty:
            return

        monthly = df.groupby("月份").agg({"總價": "sum", "利潤": "sum"}).sort_index()

        self.ax_line.plot(monthly.index, monthly["總價"], 'o-',
                          color=COLORS["primary"], linewidth=2, markersize=6,
                          label="總價")
        self.ax_line.plot(monthly.index, monthly["利潤"], 's--',
                          color=COLORS["success"], linewidth=2, markersize=6,
                          label="利潤")

        self.ax_line.set_title("月銷售額趨勢", fontsize=12, fontweight='bold', pad=10)
        self.ax_line.set_xlabel("月份")
        self.ax_line.set_ylabel("金額")
        self.ax_line.legend()
        self.ax_line.tick_params(axis='x', rotation=45, labelsize=9)

        # Y 軸千分位格式
        self.ax_line.yaxis.set_major_formatter(
            plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))

    # ─── 匯出報表 ───

    def export_report(self):
        df = self._prepare_df()
        if df.empty:
            messagebox.showwarning("提示", "沒有資料可以匯出")
            return

        filtered = self._filter_by_range(df)
        if filtered.empty:
            messagebox.showwarning("提示", "篩選範圍內沒有資料")
            return

        start = self.start_month.get() or "all"
        end = self.end_month.get() or "all"
        default_name = f"銷售報表_{start}_{end}.xlsx"

        path = filedialog.asksaveasfilename(
            title="儲存報表",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")])
        if not path:
            return

        self._write_report(filtered, path)
        messagebox.showinfo("成功", f"報表已匯出至\n{path}")

    def _write_report(self, df, path):
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        # 準備匯出欄位
        export_cols = [c for c in COLS if c in df.columns]
        export_df = df[export_cols].copy()

        # 合計行
        summary = {}
        for col in export_cols:
            if col in ["總價", "利潤", "實拿", "成本(品項+贈品)"]:
                summary[col] = pd.to_numeric(export_df[col], errors='coerce').sum()
            else:
                summary[col] = ""
        summary["日期"] = "合計"

        # 利潤率
        total_rev = summary.get("總價", 0)
        total_profit = summary.get("利潤", 0)
        profit_rate = (total_profit / total_rev * 100) if total_rev > 0 else 0

        rate_row = {c: "" for c in export_cols}
        rate_row["日期"] = "利潤率"
        rate_row["利潤"] = f"{profit_rate:.1f}%"

        export_df = pd.concat([export_df,
                                pd.DataFrame([summary]),
                                pd.DataFrame([rate_row])],
                               ignore_index=True)

        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='銷售報表')
            ws = writer.sheets['銷售報表']

            # 格式化
            header_fill = PatternFill(start_color='4A6FA5', end_color='4A6FA5', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 合計行粗體
            total_row = ws.max_row - 1
            for cell in ws[total_row]:
                cell.font = Font(bold=True, size=11)
                cell.border = thin_border

            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True, size=11)
                cell.border = thin_border

            # 自動欄寬
            for col_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = max(max_len + 4, 10)
